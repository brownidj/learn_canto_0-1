"""Assign poss_cat categories to Andy spreadsheet rows.

Scans data/andy_ss/*.xlsx and fills/updates the `poss_cat` column for each row.

Column mapping (Excel letters):
- B: Jyutping
- C: Meaning / gloss (most important for category choice)
- F: Hanzi

Rules:
- Skip header row (row 1)
- Write a comma-delimited list into poss_cat when multiple categories apply
- Prefer existing categories from data/categories.yaml
- Conservative: no new categories are created in this iteration

Design:
- UI-free
- Deterministic heuristics first (fast)
- Optional LLM assigner (OpenAI) as a plug-in; if unavailable, falls back to unassigned
- Can export proposed categories for later merging (e.g., by utils/expand_categories.py).

Usage examples (zsh):
  python3 -m utils.add_category_data --dry-run
  python3 -m utils.add_category_data --overwrite
  python3 -m utils.add_category_data --provider openai --model gpt-4.1-mini --preflight --overwrite
  python3 -m utils.add_category_data --provider openai --preflight --timeout-s 30 --max-retries 3 --backoff-s 1 --throttle-s 0.2 --overwrite
  python3 -m utils.add_category_data --propose-new-categories --propose-provider openai --export-proposed-categories data/proposed_categories.yaml
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import time
import zipfile
from dataclasses import dataclass
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from datetime import datetime

try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:  # pragma: no cover
    ZoneInfo = None  # type: ignore

try:
    from reportlab.lib.pagesizes import A4  # type: ignore
    from reportlab.pdfgen import canvas  # type: ignore
except Exception:  # pragma: no cover
    A4 = None  # type: ignore
    canvas = None  # type: ignore

import yaml

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except Exception as e:  # pragma: no cover
    raise RuntimeError("openpyxl is required for this utility: {0}".format(e))


# ----------------------------
# Paths and loading utilities
# ----------------------------


def _project_root() -> Path:
    # utils/add_category_data.py -> project root
    return Path(__file__).resolve().parent.parent


def _load_yaml(path: Path) -> object:
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def _save_yaml(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        yaml.safe_dump(payload, fh, allow_unicode=True, sort_keys=True)


def _norm_cell_text(v: object) -> str:
    if v is None:
        return ""
    try:
        s = str(v)
    except Exception:
        return ""
    return s.strip()

def _display_hanzi(hz: object) -> str:
    """Return a display-friendly Hanzi string.

    Some spreadsheets contain internal placeholder IDs like \"=D18\" in the Hanzi column.
    These are not Hanzi and should not be shown in reports.
    """
    s = _norm_cell_text(hz)
    if s.startswith("="):
        return ""
    return s


def _format_row_triplet(hz: object, jy: object, mn: object) -> str:
    """Format (hanzi?, jyutping, meaning) for console/PDF output.

    If Hanzi is missing (or is an internal placeholder ID), omit it from the display.
    """
    hz_s = _display_hanzi(hz)
    jy_s = _norm_cell_text(jy)
    mn_s = _norm_cell_text(mn)

    parts: list[str] = []
    if hz_s:
        parts.append(hz_s)
    if jy_s:
        parts.append(jy_s)
    if mn_s:
        parts.append(mn_s)

    return " | ".join(parts)

def _sha1_key(parts: Sequence[str]) -> str:
    h = hashlib.sha1()
    for p in parts:
        try:
            h.update(p.encode("utf-8"))
        except Exception:
            h.update(b"")
        h.update(b"\x1f")
    return h.hexdigest()


# ----------------------------
# Category model
# ----------------------------


@dataclass(frozen=True)
class RowPayload:
    jyutping: str
    meaning: str
    hanzi: str


# --- Unassigned row tracking and category suggestion model ---

@dataclass(frozen=True)
class UnassignedRow:
    workbook: str
    sheet: str
    row_index: int
    jyutping: str
    meaning: str
    hanzi: str


def _load_categories_map(categories_path: Path) -> Dict[str, List[str]]:
    raw = _load_yaml(categories_path)
    if not isinstance(raw, dict):
        return {}

    out: Dict[str, List[str]] = {}
    for k, v in raw.items():
        name = _norm_cell_text(k)
        if not name:
            continue
        items: List[str] = []
        if isinstance(v, list):
            for it in v:
                s = _norm_cell_text(it)
                if s:
                    items.append(s)
        out[name] = items
    return out


def _reverse_lookup_categories(categories_map: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """Build hanzi -> [categories...] map."""
    rev: Dict[str, List[str]] = {}
    for cat, items in (categories_map or {}).items():
        for hz in items:
            if not hz:
                continue
            if hz not in rev:
                rev[hz] = []
            if cat not in rev[hz]:
                rev[hz].append(cat)
    return rev


def _allowed_categories(categories_map: Dict[str, List[str]]) -> List[str]:
    cats = [c for c in (categories_map or {}).keys() if _norm_cell_text(c)]
    # Ensure unassigned is always allowed as fallback
    if "unassigned" not in cats:
        cats.append("unassigned")
    return sorted(cats)


def _dedupe_preserve_order(items: Sequence[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for it in items:
        s = _norm_cell_text(it)
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


# ----------------------------
# Assigner interface
# ----------------------------


class CategoryAssigner:
    """Assign categories for a row.

    Returns a list of category names (must be in allowed_categories) or ["unassigned"].
    """

    def assign(self, row: RowPayload, allowed_categories: Sequence[str]) -> List[str]:
        raise NotImplementedError


class NullAssigner(CategoryAssigner):
    def assign(self, row: RowPayload, allowed_categories: Sequence[str]) -> List[str]:
        _ = row
        _ = allowed_categories
        return ["unassigned"]


class OpenAIChatAssigner(CategoryAssigner):
    """OpenAI-backed assigner.

    This is intentionally optional. If `openai` is not installed, construction fails
    with a clear error.

    Environment:
    - OPENAI_API_KEY must be set

    Notes:
    - We keep the response contract strict: a JSON list of category strings.
    """

    def __init__(
            self,
            model: str,
            *,
            timeout_s: float = 30.0,
            max_retries: int = 3,
            backoff_s: float = 1.0,
            throttle_s: float = 0.0,
    ):
        self._model = _norm_cell_text(model) or "gpt-4.1-mini"
        try:
            import openai  # type: ignore

            self._openai = openai
        except Exception as e:
            raise RuntimeError(
                "OpenAI provider selected but `openai` package is not available: {0}".format(e)
            )

        if not os.environ.get("OPENAI_API_KEY"):
            raise RuntimeError("OPENAI_API_KEY is not set in the environment")

        try:
            self._timeout_s = float(timeout_s)
        except Exception:
            self._timeout_s = 30.0
        if self._timeout_s <= 0:
            self._timeout_s = 30.0

        try:
            self._max_retries = int(max_retries)
        except Exception:
            self._max_retries = 3
        if self._max_retries < 0:
            self._max_retries = 0

        try:
            self._backoff_s = float(backoff_s)
        except Exception:
            self._backoff_s = 1.0
        if self._backoff_s < 0:
            self._backoff_s = 0.0

        try:
            self._throttle_s = float(throttle_s)
        except Exception:
            self._throttle_s = 0.0
        if self._throttle_s < 0:
            self._throttle_s = 0.0

        # Build a single client with a strict httpx timeout so calls cannot hang indefinitely.
        # We keep this best-effort: if httpx is unavailable or the installed openai package
        # does not support passing a client/timeout here, we fall back to per-request timeouts.
        self._client = None
        try:
            import httpx  # type: ignore

            t = float(self._timeout_s)
            if t <= 0:
                t = 30.0

            timeout = httpx.Timeout(timeout=t, connect=t, read=t, write=t, pool=t)
            limits = httpx.Limits(max_keepalive_connections=5, max_connections=10)

            # Prefer the newer OpenAI() client.
            try:
                self._client = self._openai.OpenAI(timeout=timeout, max_retries=0)  # type: ignore
            except Exception:
                self._client = None

            # If we did not get an OpenAI client, try passing an httpx.Client explicitly.
            if self._client is None:
                try:
                    transport_client = httpx.Client(timeout=timeout, limits=limits, follow_redirects=True)
                    self._client = self._openai.OpenAI(http_client=transport_client, max_retries=0)  # type: ignore
                except Exception:
                    self._client = None
        except Exception:
            self._client = None

    def preflight(self) -> Tuple[bool, str]:
        """Verify we can successfully call the Chat Completions API.

        Returns (ok, message). This is a very small request with a strict timeout.
        """
        sys_msg = "You are a connectivity test. Reply with the single word: OK"
        user_msg = "ping"

        try:
            # Prefer newer client style if present.
            try:
                client = self._client
                if client is None:
                    client = self._openai.OpenAI()  # type: ignore

                _ = client.chat.completions.create(
                    model=self._model,
                    messages=[
                        {"role": "system", "content": sys_msg},
                        {"role": "user", "content": user_msg},
                    ],
                    temperature=0,
                    max_tokens=3,
                )
                return True, "OpenAI preflight OK (client OpenAI())."
            except Exception:
                _ = self._openai.ChatCompletion.create(
                    model=self._model,
                    messages=[
                        {"role": "system", "content": sys_msg},
                        {"role": "user", "content": user_msg},
                    ],
                    temperature=0,
                    max_tokens=3,
                    timeout=self._timeout_s,
                )
                return True, "OpenAI preflight OK (ChatCompletion.create)."
        except Exception as e:
            return False, "OpenAI preflight failed: {0}".format(e)

    def assign(self, row: RowPayload, allowed_categories: Sequence[str]) -> List[str]:
        # Keep prompt deterministic and Cantonese-focused; meaning/gloss is primary.
        allowed = [str(x) for x in allowed_categories if _norm_cell_text(x)]

        sys_msg = (
            "You are categorising colloquial Cantonese vocabulary entries. "
            "Use the provided existing categories. "
            "Return ONLY valid JSON: a list of 1+ category strings from the allowed set, "
            "or [\"unassigned\"]. "
            "If multiple categories apply, return multiple items. "
            "Do not invent new categories in this iteration."
        )

        user_msg = (
            "Entry:\n"
            "- Hanzi: {0}\n"
            "- Jyutping: {1}\n"
            "- Meaning: {2}\n\n"
            "Allowed categories:\n{3}\n\n"
            "Return JSON list only."
        ).format(row.hanzi, row.jyutping, row.meaning, json.dumps(allowed, ensure_ascii=False))

        content = None

        # Retry loop; client-level httpx timeouts prevent indefinite hangs. We also optionally throttle.
        attempt = 0
        last_err: Optional[Exception] = None
        while attempt <= self._max_retries:
            if self._throttle_s:
                try:
                    time.sleep(self._throttle_s)
                except Exception:
                    pass

            try:
                # Support both older and newer openai python client styles.
                try:
                    client = self._client
                    if client is None:
                        client = self._openai.OpenAI()  # type: ignore

                    resp = client.chat.completions.create(
                        model=self._model,
                        messages=[
                            {"role": "system", "content": sys_msg},
                            {"role": "user", "content": user_msg},
                        ],
                        temperature=0,
                    )
                    content = resp.choices[0].message.content
                except Exception:
                    resp = self._openai.ChatCompletion.create(
                        model=self._model,
                        messages=[
                            {"role": "system", "content": sys_msg},
                            {"role": "user", "content": user_msg},
                        ],
                        temperature=0,
                        timeout=self._timeout_s,
                    )
                    content = resp["choices"][0]["message"]["content"]

                # If we got here, the request succeeded.
                last_err = None
                break
            except Exception as e:
                last_err = e
                attempt += 1
                if attempt > self._max_retries:
                    break
                # Simple backoff.
                try:
                    time.sleep(self._backoff_s * float(attempt))
                except Exception:
                    pass

        if last_err is not None and not content:
            return ["unassigned"]

        try:
            parsed = json.loads(content or "")
        except Exception:
            return ["unassigned"]

        if not isinstance(parsed, list):
            return ["unassigned"]

        cleaned: List[str] = []
        for it in parsed:
            s = _norm_cell_text(it)
            if not s:
                continue
            cleaned.append(s)

        cleaned = _dedupe_preserve_order(cleaned)
        if not cleaned:
            return ["unassigned"]

        allowed_set = set([_norm_cell_text(x) for x in allowed])
        ok: List[str] = []
        for c in cleaned:
            if c in allowed_set:
                ok.append(c)

        if not ok:
            return ["unassigned"]

        return ok


class OpenAIChatCategoryProposer:
    """OpenAI-backed proposer for *new* category names (report only; does not write YAML)."""

    def __init__(
            self,
            model: str,
            *,
            timeout_s: float = 30.0,
            max_retries: int = 2,
            backoff_s: float = 1.0,
            throttle_s: float = 0.0,
    ):
        self._model = _norm_cell_text(model) or "gpt-4.1-mini"
        try:
            import openai  # type: ignore
            self._openai = openai
        except Exception as e:
            raise RuntimeError("OpenAI proposer selected but `openai` package is not available: {0}".format(e))

        if not os.environ.get("OPENAI_API_KEY"):
            raise RuntimeError("OPENAI_API_KEY is not set in the environment")

        try:
            self._timeout_s = float(timeout_s)
        except Exception:
            self._timeout_s = 30.0
        if self._timeout_s <= 0:
            self._timeout_s = 30.0

        try:
            self._max_retries = int(max_retries)
        except Exception:
            self._max_retries = 2
        if self._max_retries < 0:
            self._max_retries = 0

        try:
            self._backoff_s = float(backoff_s)
        except Exception:
            self._backoff_s = 1.0
        if self._backoff_s < 0:
            self._backoff_s = 0.0

        try:
            self._throttle_s = float(throttle_s)
        except Exception:
            self._throttle_s = 0.0
        if self._throttle_s < 0:
            self._throttle_s = 0.0

        self._client = None
        try:
            import httpx  # type: ignore
            t = float(self._timeout_s)
            if t <= 0:
                t = 30.0
            timeout = httpx.Timeout(timeout=t, connect=t, read=t, write=t, pool=t)
            try:
                self._client = self._openai.OpenAI(timeout=timeout, max_retries=0)  # type: ignore
            except Exception:
                self._client = None
        except Exception:
            self._client = None

    def preflight(self) -> Tuple[bool, str]:
        sys_msg = "You are a connectivity test. Reply with the single word: OK"
        user_msg = "ping"
        try:
            client = self._client
            if client is None:
                client = self._openai.OpenAI()  # type: ignore
            _ = client.chat.completions.create(
                model=self._model,
                messages=[
                    {"role": "system", "content": sys_msg},
                    {"role": "user", "content": user_msg},
                ],
                temperature=0,
                max_tokens=3,
            )
            return True, "OpenAI proposer preflight OK."
        except Exception as e:
            return False, "OpenAI proposer preflight failed: {0}".format(e)

    def propose(
            self,
            rows: Sequence["UnassignedRow"],
            existing_categories: Sequence[str],
            *,
            min_rows: int,
            max_categories: int,
            batch_size: int = 25,
    ) -> List[Dict[str, object]]:
        all_rows = list(rows or [])
        if not all_rows:
            return []

        try:
            min_rows_i = int(min_rows)
        except Exception:
            min_rows_i = 3
        if min_rows_i < 2:
            min_rows_i = 2

        try:
            max_cats_i = int(max_categories)
        except Exception:
            max_cats_i = 10
        if max_cats_i <= 0:
            max_cats_i = 10

        try:
            bs = int(batch_size)
        except Exception:
            bs = 25
        if bs <= 0:
            bs = 25

        existing = [str(x) for x in (existing_categories or []) if _norm_cell_text(x)]
        existing_l = set([x.lower() for x in existing])

        sys_msg = (
            "You are proposing NEW category names for colloquial Cantonese vocabulary entries. "
            "Do NOT repeat any existing category name. "
            "Only propose a category if it clearly groups at least min_rows items in the provided list. "
            "Return ONLY valid JSON: a list of objects with keys: category, rows, rationale. "
            "category must be short (snake_case or lowercase phrase). "
            "rows must be a list of 0-based indices into the provided batch. "
            "No markdown."
        )

        merged: Dict[str, Dict[str, object]] = {}

        start = 0
        while start < len(all_rows) and len(merged) < max_cats_i:
            batch = all_rows[start: start + bs]

            payload_rows: List[Dict[str, str]] = []
            for i, r in enumerate(batch):
                payload_rows.append(
                    {"i": str(i), "hanzi": r.hanzi, "jyutping": r.jyutping, "meaning": r.meaning}
                )

            user_msg = (
                "Existing categories:\n{0}\n\n"
                "min_rows={1}\n\n"
                "Unassigned rows (0-based indices):\n{2}\n\n"
                "Return JSON list only."
            ).format(
                json.dumps(existing, ensure_ascii=False),
                str(min_rows_i),
                json.dumps(payload_rows, ensure_ascii=False),
            )

            content = ""
            attempt = 0
            while attempt <= self._max_retries:
                if self._throttle_s:
                    try:
                        time.sleep(self._throttle_s)
                    except Exception:
                        pass

                try:
                    client = self._client
                    if client is None:
                        client = self._openai.OpenAI()  # type: ignore

                    resp = client.chat.completions.create(
                        model=self._model,
                        messages=[
                            {"role": "system", "content": sys_msg},
                            {"role": "user", "content": user_msg},
                        ],
                        temperature=0,
                    )
                    content = resp.choices[0].message.content or ""
                    break
                except Exception:
                    attempt += 1
                    if attempt > self._max_retries:
                        break
                    try:
                        time.sleep(self._backoff_s * float(attempt))
                    except Exception:
                        pass

            try:
                parsed = json.loads(content or "")
            except Exception:
                parsed = []

            if isinstance(parsed, list):
                for obj in parsed:
                    if not isinstance(obj, dict):
                        continue
                    name = _norm_cell_text(obj.get("category"))
                    if not name:
                        continue
                    if name.lower() in existing_l:
                        continue

                    idxs = obj.get("rows")
                    if not isinstance(idxs, list):
                        idxs = []
                    idxs_i: List[int] = []
                    for x in idxs:
                        try:
                            xi = int(x)
                        except Exception:
                            continue
                        if 0 <= xi < len(batch):
                            idxs_i.append(xi)

                    idxs_i = sorted(set(idxs_i))
                    if len(idxs_i) < min_rows_i:
                        continue

                    rationale = _norm_cell_text(obj.get("rationale"))

                    examples: List[Dict[str, str]] = []
                    for xi in idxs_i[:5]:
                        rr = batch[xi]
                        examples.append(
                            {
                                "hanzi": rr.hanzi,
                                "jyutping": rr.jyutping,
                                "meaning": rr.meaning,
                                "workbook": rr.workbook,
                                "sheet": rr.sheet,
                                "row": str(rr.row_index),
                            }
                        )

                    merged[name] = {
                        "proposed_category": name,
                        "supporting_rows": int(len(idxs_i)),
                        "examples": examples,
                        "rationale": rationale,
                    }

                    if len(merged) >= max_cats_i:
                        break

            start += bs

        out = list(merged.values())
        out = sorted(
            out,
            key=lambda d: (-int(d.get("supporting_rows") or 0), _norm_cell_text(d.get("proposed_category"))),
        )
        return out[:max_cats_i]


def _choose_category_proposer(
        provider: str,
        model: str,
        *,
        timeout_s: float,
        max_retries: int,
        backoff_s: float,
        throttle_s: float,
):
    p = _norm_cell_text(provider).lower()
    if p == "openai":
        return OpenAIChatCategoryProposer(
            model=model,
            timeout_s=timeout_s,
            max_retries=max_retries,
            backoff_s=backoff_s,
            throttle_s=throttle_s,
        )
    return None


# ----------------------------
# Excel helpers
# ----------------------------


def _find_or_create_poss_cat_col(ws: Worksheet) -> int:
    """Return 1-based column index for poss_cat, creating if missing."""
    header_row = 1
    max_col = ws.max_column or 1

    poss_col = -1
    for col in range(1, max_col + 1):
        v = _norm_cell_text(ws.cell(row=header_row, column=col).value)
        if v.lower() == "poss_cat":
            poss_col = col
            break

    if poss_col > 0:
        return poss_col

    poss_col = max_col + 1
    ws.cell(row=header_row, column=poss_col).value = "poss_cat"
    return poss_col


def _iter_data_rows(ws: Worksheet) -> Iterable[int]:
    # Openpyxl uses 1-based row indices
    max_row = ws.max_row or 1
    for r in range(2, max_row + 1):
        yield r


def _read_row_payload(ws: Worksheet, row: int) -> RowPayload:
    jy = _read_cell_text(ws, row, 2)  # B
    mn = _read_cell_text(ws, row, 3)  # C
    hz = _read_cell_text(ws, row, 6)  # F

    # If the sheet contains placeholder squares, treat as missing for reporting.
    if hz in ("■", "□"):
        hz = ""

    return RowPayload(jyutping=jy, meaning=mn, hanzi=hz)

_A1_RE = re.compile(r"^\s*=?\s*([A-Za-z]{1,3})(\d{1,7})\s*$")

def _col_letters_to_index(col_letters: str) -> int:
    s = _norm_cell_text(col_letters).upper()
    if not s:
        return 0
    n = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            return 0
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def _resolve_cell_ref_text(ws: Worksheet, text: str, *, _depth: int = 0) -> str:
    """Resolve A1-style refs like '=B63' to the target cell's text."""
    if _depth >= 5:
        return _norm_cell_text(text)

    s = _norm_cell_text(text)
    if not s:
        return ""

    m = _A1_RE.match(s)
    if not m:
        return s

    c = _col_letters_to_index(m.group(1))
    try:
        r = int(m.group(2))
    except Exception:
        return s

    if c <= 0 or r <= 0:
        return s

    try:
        target_val = ws.cell(row=r, column=c).value
    except Exception:
        return s

    t = _norm_cell_text(target_val)
    if not t:
        return ""

    if t.startswith("="):
        return _resolve_cell_ref_text(ws, t, _depth=_depth + 1)

    return t

def _read_cell_text(ws: Worksheet, row: int, col: int) -> str:
    try:
        raw = ws.cell(row=row, column=col).value
    except Exception:
        raw = None

    s = _norm_cell_text(raw)
    if not s:
        return ""

    if s.startswith("="):
        try:
            return _norm_cell_text(_resolve_cell_ref_text(ws, s))
        except Exception:
            return s

    return s


# ----------------------------
# Orchestration
# ----------------------------


@dataclass
class RunStats:
    files: int = 0
    sheets: int = 0
    rows_seen: int = 0
    rows_updated: int = 0
    rows_skipped_existing: int = 0
    rows_skipped_empty: int = 0
    rows_unassigned: int = 0
    rows_assigned: int = 0
    last_progress_ts: float = 0.0
    unassigned_rows: List['UnassignedRow'] = None  # filled lazily
    rows_recategorise_seen: int = 0
    rows_recategorise_changed: int = 0
    rows_recategorise_skipped: int = 0
    rows_recategorise_still_unassigned: int = 0
    rows_recategorise_became_assigned: int = 0


# --- Heuristic category proposal helpers ---

_STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "be", "but", "by", "for", "from", "has", "have", "he", "her",
    "his", "i", "in", "is", "it", "its", "me", "my", "of", "on", "or", "our", "she", "so", "that",
    "the", "their", "them", "then", "there", "these", "they", "this", "to", "was", "we", "were", "with",
    "you", "your",
}


def _safe_init_unassigned_list(stats: RunStats) -> None:
    try:
        if stats.unassigned_rows is None:
            stats.unassigned_rows = []
    except Exception:
        try:
            stats.unassigned_rows = []
        except Exception:
            pass


def _tokenize_meaning(text: str) -> List[str]:
    s = _norm_cell_text(text).lower()
    if not s:
        return []
    # Keep simple latin tokens; this is heuristic only.
    parts = re.split(r"[^a-z0-9']+", s)
    out: List[str] = []
    for p in parts:
        p2 = _norm_cell_text(p)
        if not p2:
            continue
        if len(p2) < 3:
            continue
        if p2 in _STOPWORDS:
            continue
        out.append(p2)
    return out


def _propose_new_categories_heuristic(
        unassigned_rows: Sequence['UnassignedRow'],
        existing_categories: Sequence[str],
        *,
        min_rows: int = 3,
        max_categories: int = 10,
        examples_per_cat: int = 5,
) -> List[dict]:
    """Heuristic proposer: cluster unassigned rows by common gloss tokens.

    This does NOT write any new categories; it only suggests candidates.
    """
    rows = list(unassigned_rows or [])
    if not rows:
        return []

    existing = set([_norm_cell_text(x).lower() for x in (existing_categories or []) if _norm_cell_text(x)])

    token_to_rows: Dict[str, List[int]] = {}
    token_counts: Dict[str, int] = {}

    for idx, r in enumerate(rows):
        toks = _tokenize_meaning(r.meaning)
        # Deduplicate tokens per row to avoid overweighting repeated words.
        seen = set()
        for t in toks:
            if t in seen:
                continue
            seen.add(t)
            token_counts[t] = int(token_counts.get(t, 0)) + 1
            token_to_rows.setdefault(t, []).append(idx)

    # Rank tokens by count (desc), then alphabetically for determinism.
    ranked = sorted(token_counts.items(), key=lambda kv: (-int(kv[1]), str(kv[0])))

    proposals: List[Dict[str, object]] = []
    used_row_indices: set[int] = set()

    for token, cnt in ranked:
        if int(cnt) < int(min_rows):
            continue
        if token in existing:
            continue

        idxs = token_to_rows.get(token) or []
        # Filter out rows already captured by earlier proposals to keep the report concise.
        idxs2 = [i for i in idxs if i not in used_row_indices]
        if len(idxs2) < int(min_rows):
            continue

        ex: List[Dict[str, str]] = []
        for i in idxs2[: int(examples_per_cat)]:
            rr = rows[i]
            ex.append(
                {
                    "hanzi": rr.hanzi,
                    "jyutping": rr.jyutping,
                    "meaning": rr.meaning,
                    "workbook": rr.workbook,
                    "sheet": rr.sheet,
                    "row": str(rr.row_index),
                }
            )

        proposals.append(
            {
                "proposed_category": token,
                "supporting_rows": int(len(idxs2)),
                "examples": ex,
            }
        )

        for i in idxs2:
            used_row_indices.add(i)

        if len(proposals) >= int(max_categories):
            break

    return proposals


def _print_proposed_categories_report(
        proposals: Sequence[dict],
        *,
        prefix: str = "",
) -> None:
    props = list(proposals or [])
    if not props:
        print(prefix + "No new category suggestions.")
        return

    print(prefix + "Proposed new categories (report only; not written):")
    for p in props:
        name = _norm_cell_text(p.get("proposed_category"))
        n = p.get("supporting_rows")
        try:
            n_int = int(n)
        except Exception:
            n_int = 0
        rationale = _norm_cell_text(p.get("rationale"))
        if rationale:
            print(prefix + "- {0} (rows≈{1}) — {2}".format(name or "(unnamed)", n_int, rationale))
        else:
            print(prefix + "- {0} (rows≈{1})".format(name or "(unnamed)", n_int))

        ex = p.get("examples")
        if isinstance(ex, list) and ex:
            for e in ex[:5]:
                if not isinstance(e, dict):
                    continue
                hz = _norm_cell_text(e.get("hanzi"))
                jy = _norm_cell_text(e.get("jyutping"))
                mn = _norm_cell_text(e.get("meaning"))
                wb = _norm_cell_text(e.get("workbook"))
                sh = _norm_cell_text(e.get("sheet"))
                rw = _norm_cell_text(e.get("row"))
                print(prefix + "    • {0} | {1} | {2} ({3}:{4} r{5})".format(hz, jy, mn, wb, sh, rw))


def _today_ymd_brisbane() -> str:
    """Return today's date in Australia/Brisbane as YYYY-MM-DD."""
    try:
        if ZoneInfo is not None:
            dt = datetime.now(ZoneInfo("Australia/Brisbane"))
        else:
            dt = datetime.now()
    except Exception:
        dt = datetime.now()
    try:
        return dt.date().isoformat()
    except Exception:
        return datetime.now().date().isoformat()


def _wrap_text_to_width(c, text: str, *, max_width: float, font_name: str, font_size: int) -> List[str]:
    """Simple word-wrapping for ReportLab canvas text."""
    try:
        words = (text or "").split()
    except Exception:
        words = [str(text)]
    if not words:
        return [""]

    lines: List[str] = []
    cur: List[str] = []
    for w in words:
        trial = (" ".join(cur + [w])).strip()
        try:
            tw = c.stringWidth(trial, font_name, font_size)
        except Exception:
            tw = 0
        if cur and tw > max_width:
            lines.append(" ".join(cur))
            cur = [w]
        else:
            cur.append(w)

    if cur:
        lines.append(" ".join(cur))
    return lines


def _write_pdf_report_for_workbook(
        xlsx_path: Path,
        *,
        root_path: Path,
        stats: 'RunStats',
        unassigned_rows: Sequence['UnassignedRow'],
        categories: Sequence[str],
        dry_run: bool,
) -> None:
    """Write a per-workbook categorisation report PDF alongside the workbook."""
    if dry_run:
        return

    if canvas is None or A4 is None:
        print("WARNING: reportlab not available; skipping PDF report for {0}".format(xlsx_path))
        return

    date_s = _today_ymd_brisbane()
    try:
        stem = xlsx_path.stem
    except Exception:
        stem = "workbook"

    # NOTE: spelling per your request: catagorisation
    out_pdf = xlsx_path.with_name("{0}_catagorisation_{1}.pdf".format(stem, date_s))

    try:
        rel = xlsx_path.relative_to(root_path)
        workbook_label = str(rel)
    except Exception:
        workbook_label = str(xlsx_path)

    c = canvas.Canvas(str(out_pdf), pagesize=A4)
    width, height = A4

    margin_x = 48
    y = height - 54
    line_h = 14
    font_body = "Helvetica"
    font_bold = "Helvetica-Bold"

    def draw_line(txt: str, *, bold: bool = False) -> None:
        nonlocal y
        if y < 54:
            c.showPage()
            y = height - 54
        c.setFont(font_bold if bold else font_body, 11)
        c.drawString(margin_x, y, txt)
        y -= line_h

    def draw_wrapped(label: str, value: str) -> None:
        nonlocal y
        txt = "{0}{1}".format(label, value)
        c.setFont(font_body, 11)
        max_w = width - (2 * margin_x)
        lines = _wrap_text_to_width(c, txt, max_width=max_w, font_name=font_body, font_size=11)
        for ln in lines:
            if y < 54:
                c.showPage()
                y = height - 54
            c.drawString(margin_x, y, ln)
            y -= line_h

    draw_line("Categorisation report", bold=True)
    draw_line("Workbook: {0}".format(workbook_label))
    draw_line("Date: {0}".format(date_s))
    draw_line("")

    draw_line("Summary", bold=True)
    draw_wrapped("  sheets: ", str(getattr(stats, "sheets", 0)))
    draw_wrapped("  rows_seen: ", str(getattr(stats, "rows_seen", 0)))
    draw_wrapped("  updated: ", str(getattr(stats, "rows_updated", 0)))
    draw_wrapped("  skipped_existing: ", str(getattr(stats, "rows_skipped_existing", 0)))
    draw_wrapped("  empty: ", str(getattr(stats, "rows_skipped_empty", 0)))
    draw_wrapped("  assigned: ", str(getattr(stats, "rows_assigned", 0)))
    draw_wrapped("  unassigned: ", str(getattr(stats, "rows_unassigned", 0)))
    draw_line("")

    rec_seen = int(getattr(stats, "rows_recategorise_seen", 0) or 0)
    if rec_seen:
        draw_line("Recategorisation", bold=True)
        draw_wrapped("  rows_recategorise_seen: ", str(rec_seen))
        draw_wrapped("  rows_recategorise_changed: ", str(getattr(stats, "rows_recategorise_changed", 0)))
        draw_wrapped("  rows_recategorise_skipped: ", str(getattr(stats, "rows_recategorise_skipped", 0)))
        draw_wrapped("  still_unassigned: ", str(getattr(stats, "rows_recategorise_still_unassigned", 0)))
        draw_wrapped("  became_assigned: ", str(getattr(stats, "rows_recategorise_became_assigned", 0)))
        draw_line("")

    rows_u = list(unassigned_rows or [])
    if rows_u:
        draw_line("Unassigned rows", bold=True)

        miss_hz = miss_jy = miss_mn = 0
        for r in rows_u:
            if not _norm_cell_text(getattr(r, "hanzi", "")):
                miss_hz += 1
            if not _norm_cell_text(getattr(r, "jyutping", "")):
                miss_jy += 1
            if not _norm_cell_text(getattr(r, "meaning", "")):
                miss_mn += 1

        draw_wrapped(
            "  missing hanzi/jyutping/meaning: ",
            "{0}/{1}, {2}/{1}, {3}/{1}".format(miss_hz, len(rows_u), miss_jy, miss_mn),
        )
        draw_line("")

        token_counts: Dict[str, int] = {}
        for r in rows_u:
            for t in _tokenize_meaning(getattr(r, "meaning", "")):
                token_counts[t] = token_counts.get(t, 0) + 1

        top_tokens = sorted(token_counts.items(), key=lambda kv: (-kv[1], kv[0]))[:10]
        if top_tokens:
            draw_line("  Top meaning tokens:", bold=True)
            for tok, cnt in top_tokens:
                draw_wrapped("    - ", "{0}: {1}".format(tok, cnt))
            draw_line("")

        draw_line("  Rows:", bold=True)
        for r in rows_u:
            hz = _norm_cell_text(getattr(r, "hanzi", ""))
            jy = _norm_cell_text(getattr(r, "jyutping", ""))
            mn = _norm_cell_text(getattr(r, "meaning", ""))
            wb = _norm_cell_text(getattr(r, "workbook", ""))
            sh = _norm_cell_text(getattr(r, "sheet", ""))
            try:
                ri = int(getattr(r, "row_index", 0))
            except Exception:
                ri = 0
            draw_wrapped("    • ", "{0} ({1}:{2} r{3})".format(_format_row_triplet(hz, jy, mn), wb, sh, ri).format(hz, jy, mn, wb, sh, ri))
    else:
        draw_line("Unassigned rows: none.", bold=True)

    # --- Categories list (alphabetical) ---
    try:
        cats = [
            _norm_cell_text(x)
            for x in (categories or [])
            if _norm_cell_text(x)
        ]
    except Exception:
        cats = []

    # Dedupe + sort (case-insensitive)
    try:
        cats = sorted(set(cats), key=lambda s: s.lower())
    except Exception:
        try:
            cats = sorted(set(cats))
        except Exception:
            pass

    draw_line("")
    draw_line("Categories (alphabetical)", bold=True)
    if not cats:
        draw_line("(none)")
    else:
        for cat_name in cats:
            draw_wrapped("  - ", cat_name)
    c.save()
    print("Wrote report: {0}".format(out_pdf))

# --- End-of-run unassigned analysis report ---
def _print_unassigned_analysis(unassigned: Sequence['UnassignedRow'], *, max_rows: Optional[int] = 10) -> None:
    if not unassigned:
        print("Unassigned analysis: no unassigned rows remain.")
        return

    print("Unassigned analysis:")

    # Missing-field breakdown
    miss_hz = miss_jy = miss_mn = 0
    for r in unassigned:
        if not _norm_cell_text(r.hanzi):
            miss_hz += 1
        if not _norm_cell_text(r.jyutping):
            miss_jy += 1
        if not _norm_cell_text(r.meaning):
            miss_mn += 1

    total = len(unassigned)
    print(
        "  Missing fields: hanzi={0}/{1}, jyutping={2}/{1}, meaning={3}/{1}".format(
            miss_hz, total, miss_jy, miss_mn
        )
    )

    # Token frequency in meanings
    token_counts: Dict[str, int] = {}
    for r in unassigned:
        for t in _tokenize_meaning(r.meaning):
            token_counts[t] = token_counts.get(t, 0) + 1

    top_tokens = sorted(token_counts.items(), key=lambda kv: (-kv[1], kv[0]))[:10]
    if top_tokens:
        print("  Top meaning tokens:")
        for tok, cnt in top_tokens:
            print("    - {0}: {1}".format(tok, cnt))

    # Sample rows
    print("  Sample unassigned rows:")
    try:
        n = max_rows
    except Exception:
        n = 10

    if n is None:
        print("  Unassigned rows (all):")
        seq = list(unassigned)
    else:
        try:
            n_int = int(n)
        except Exception:
            n_int = 10
        if n_int < 0:
            n_int = 0
        print("  Unassigned rows (showing up to {0}):".format(n_int))
        seq = list(unassigned)[:n_int]

    for r in seq:
        print(
            "    • {0} ({1}:{2} r{3})".format(
                _format_row_triplet(r.hanzi, r.jyutping, r.meaning),
                r.workbook,
                r.sheet,
                r.row_index,
            )
        )


def _print_unassigned_pattern_summary_from_proposals(
        proposals: Sequence[dict],
        unassigned: Sequence['UnassignedRow'],
        *,
        prefix: str = "",
) -> None:
    """Concrete pattern summary: show examples under each proposed category, then list the rest."""
    props = list(proposals or [])
    rows = list(unassigned or [])

    if not rows:
        print(prefix + "Unassigned pattern summary: no unassigned rows remain.")
        return

    if not props:
        print(prefix + "Unassigned pattern summary: no proposals available.")
        return

    # Lookup unassigned rows by (workbook, sheet, row)
    key_to_row: Dict[Tuple[str, str, int], 'UnassignedRow'] = {}
    for r in rows:
        try:
            key = (_norm_cell_text(r.workbook), _norm_cell_text(r.sheet), int(r.row_index))
        except Exception:
            continue
        key_to_row[key] = r

    covered: set[Tuple[str, str, int]] = set()

    print(prefix + "Unassigned pattern summary (from proposals):")
    for p in props:
        name = _norm_cell_text(p.get("proposed_category")) or "(unnamed)"
        try:
            n_int = int(p.get("supporting_rows") or 0)
        except Exception:
            n_int = 0
        rationale = _norm_cell_text(p.get("rationale"))

        if rationale:
            print(prefix + "  Pattern: {0} (≈{1} rows) — {2}".format(name, n_int, rationale))
        else:
            print(prefix + "  Pattern: {0} (≈{1} rows)".format(name, n_int))

        ex = p.get("examples")
        any_printed = False
        if isinstance(ex, list) and ex:
            for e in ex:
                if not isinstance(e, dict):
                    continue
                wb = _norm_cell_text(e.get("workbook"))
                sh = _norm_cell_text(e.get("sheet"))
                rw = _norm_cell_text(e.get("row"))
                try:
                    ri = int(rw)
                except Exception:
                    continue

                key = (wb, sh, ri)
                rr = key_to_row.get(key)
                if rr is None:
                    continue

                covered.add(key)
                any_printed = True
                print(
                    prefix + "    • {0} ({1}:{2} r{3})".format(
                        _format_row_triplet(rr.hanzi, rr.jyutping, rr.meaning),
                        # or r.hanzi/r.jyutping/r.meaning in the remaining loop
                        rr.workbook,  # or r.workbook
                        rr.sheet,  # or r.sheet
                        rr.row_index,  # or r.row_index
                    )
                )

        if not any_printed:
            print(prefix + "    • (no matching unassigned example rows available)")

    remaining: List['UnassignedRow'] = []
    for r in rows:
        try:
            key = (_norm_cell_text(r.workbook), _norm_cell_text(r.sheet), int(r.row_index))
        except Exception:
            continue
        if key not in covered:
            remaining.append(r)

    if remaining:
        print(prefix + "  Remaining unassigned rows (not covered by proposal examples): {0}".format(len(remaining)))
        for r in remaining:
            print(
                prefix + "    • {0} | {1} | {2} ({3}:{4} r{5})".format(
                    r.hanzi,
                    r.jyutping,
                    r.meaning,
                    r.workbook,
                    r.sheet,
                    r.row_index,
                )
            )


# --- Export proposed categories to YAML ---
def _export_proposed_categories_yaml(
        proposals: Sequence[dict],
        out_path: Path,
        *,
        include_example_hanzi: bool = False,
) -> None:
    """Write proposed categories to YAML in a categories.yaml-compatible structure.

    Output shape:
      {"category_name": ["漢字1", "漢字2", ...], ...}

    By default, lists are empty (safe). If include_example_hanzi=True, we seed each category
    with the distinct example Hanzi strings shown in the report.
    """
    props = list(proposals or [])
    payload: Dict[str, List[str]] = {}

    for p in props:
        name = _norm_cell_text(p.get("proposed_category"))
        if not name:
            continue

        if not include_example_hanzi:
            payload[name] = []
            continue

        hz_list: List[str] = []
        ex = p.get("examples")
        if isinstance(ex, list):
            for e in ex:
                if not isinstance(e, dict):
                    continue
                hz = _norm_cell_text(e.get("hanzi"))
                if hz:
                    hz_list.append(hz)

        payload[name] = _dedupe_preserve_order(hz_list)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as fh:
        yaml.safe_dump(payload, fh, allow_unicode=True, sort_keys=True)


def _merge_proposed_categories_into_categories_yaml(
        proposals: Sequence[dict],
        categories_yaml_path: Path,
        *,
        seed_example_hanzi: bool = False,
) -> Tuple[int, int]:
    """Merge proposed categories into an existing categories.yaml.

    Preserves formatting/comments/order using ruamel.yaml round-trip YAML.

    Returns: (added_categories, skipped_existing).
    """
    props = list(proposals or [])
    if not props:
        return 0, 0

    try:
        from ruamel.yaml import YAML  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "ruamel.yaml is required for --apply-proposed-categories (pip install ruamel.yaml): {0}".format(e)
        )

    yaml_rt = YAML()
    yaml_rt.preserve_quotes = True

    if not categories_yaml_path.exists():
        raise FileNotFoundError("categories.yaml not found: {0}".format(categories_yaml_path))

    with open(categories_yaml_path, "r", encoding="utf-8") as fh:
        doc = yaml_rt.load(fh)

    # categories.yaml is expected to be a mapping of category -> list OR category -> {items: [...]}
    if doc is None:
        doc = {}

    if not isinstance(doc, dict):
        raise TypeError("categories.yaml must be a mapping, got: {0}".format(type(doc)))

    # Detect schema: if any existing category value is a dict containing 'items', use that for new cats.
    use_items_schema = False
    try:
        for _k, _v in doc.items():
            if isinstance(_v, dict) and "items" in _v:
                use_items_schema = True
                break
    except Exception:
        use_items_schema = False

    added = 0
    skipped = 0

    for p in props:
        name = _norm_cell_text(p.get("proposed_category"))
        if not name:
            continue

        if name in doc:
            skipped += 1
            continue

        seed: List[str] = []
        if seed_example_hanzi:
            ex = p.get("examples")
            if isinstance(ex, list):
                for e in ex:
                    if not isinstance(e, dict):
                        continue
                    hz = _norm_cell_text(e.get("hanzi"))
                    if hz:
                        seed.append(hz)
            seed = _dedupe_preserve_order(seed)

        if use_items_schema:
            doc[name] = {"items": list(seed)}
        else:
            doc[name] = list(seed)

        added += 1

    with open(categories_yaml_path, "w", encoding="utf-8") as fh:
        yaml_rt.dump(doc, fh)

    return added, skipped


def _choose_assigner(
        provider: str,
        model: str,
        *,
        timeout_s: float,
        max_retries: int,
        backoff_s: float,
        throttle_s: float,
) -> CategoryAssigner:
    p = _norm_cell_text(provider).lower()
    if p == "openai":
        return OpenAIChatAssigner(
            model=model,
            timeout_s=timeout_s,
            max_retries=max_retries,
            backoff_s=backoff_s,
            throttle_s=throttle_s,
        )
    return NullAssigner()


def _assign_for_row(
        row: RowPayload,
        *,
        allowed: Sequence[str],
        hanzi_to_cats: Dict[str, List[str]],
        assigner: CategoryAssigner,
        cache: Dict[str, object],
) -> List[str]:
    # Deterministic: if hanzi already appears in categories.yaml, use those.
    hz = _norm_cell_text(row.hanzi)
    if hz and hz in (hanzi_to_cats or {}):
        cats = hanzi_to_cats.get(hz) or []
        cats = [c for c in cats if c in set(allowed)]
        cats = _dedupe_preserve_order(cats)
        if cats:
            return cats

    key = _sha1_key([row.hanzi, row.jyutping, row.meaning])
    if key in cache:
        cached = cache.get(key)
        if isinstance(cached, list):
            cached_list = [str(x) for x in cached if _norm_cell_text(x)]
            cached_list = _dedupe_preserve_order(cached_list)
            if cached_list:
                return cached_list

    cats = assigner.assign(row, allowed)
    cats = _dedupe_preserve_order(cats)
    if not cats:
        cats = ["unassigned"]

    # Constrain to allowed (fallback to unassigned)
    allowed_set = set([_norm_cell_text(x) for x in allowed])
    pruned = [c for c in cats if c in allowed_set]
    if not pruned:
        pruned = ["unassigned"]

    cache[key] = pruned
    return pruned


def _parse_poss_cat_cell(val: object) -> List[str]:
    try:
        s = _norm_cell_text(val)
    except Exception:
        s = ""
    if not s:
        return []
    try:
        parts = [p.strip() for p in str(s).split(",") if p.strip()]
    except Exception:
        parts = [str(s).strip()] if str(s).strip() else []
    out: List[str] = []
    for p in parts:
        p2 = _norm_cell_text(p)
        if p2:
            out.append(p2)
    return _dedupe_preserve_order(out)


def _format_poss_cat_cell(cats: Sequence[str]) -> str:
    seq = [_norm_cell_text(x) for x in (cats or []) if _norm_cell_text(x)]
    seq = _dedupe_preserve_order(seq)
    if not seq:
        return ""
    return ", ".join(seq)


def _merge_categories(existing: Sequence[str], proposed: Sequence[str], *, policy: str) -> List[str]:
    ex = [_norm_cell_text(x) for x in (existing or []) if _norm_cell_text(x)]
    pr = [_norm_cell_text(x) for x in (proposed or []) if _norm_cell_text(x)]
    ex = _dedupe_preserve_order(ex)
    pr = _dedupe_preserve_order(pr)

    pol = _norm_cell_text(policy).lower()
    if pol not in ("append", "replace"):
        pol = "append"

    if pol == "replace":
        out = pr
    else:
        out = _dedupe_preserve_order(list(ex) + list(pr))

    out_l = [x.lower() for x in out]
    if "unassigned" in out_l and len(out_l) > 1:
        out2: List[str] = []
        for x in out:
            if _norm_cell_text(x).lower() == "unassigned":
                continue
            out2.append(x)
        out = out2

    if not out:
        out = ["unassigned"]

    return out


def process_workbook(
        xlsx_path: Path,
        *,
        categories_map: Dict[str, List[str]],
        assigner: CategoryAssigner,
        cache: Dict[str, object],
        dry_run: bool,
        overwrite: bool,
        force: bool,
        limit: Optional[int],
        workbook_label: Optional[str] = None,
        collect_unassigned_from_existing: bool = False,
        recategorise: bool = False,
        recategorise_scope: str = "unassigned",
        recategorise_policy: str = "append",
) -> Tuple[bool, RunStats]:
    stats = RunStats(files=1)
    _safe_init_unassigned_list(stats)

    # Normalize recategorise options
    recat_enabled = bool(recategorise)
    recat_scope = _norm_cell_text(recategorise_scope).lower() or "unassigned"
    if recat_scope not in ("unassigned", "all"):
        recat_scope = "unassigned"

    recat_policy = _norm_cell_text(recategorise_policy).lower() or "append"
    if recat_policy not in ("append", "replace"):
        recat_policy = "append"

    progress_interval = 5.0  # seconds between progress messages

    allowed = _allowed_categories(categories_map)
    hanzi_to_cats = _reverse_lookup_categories(categories_map)

    try:
        wb: Workbook = load_workbook(filename=str(xlsx_path))
    except (zipfile.BadZipFile, InvalidFileException) as e:
        print(
            "WARNING: skipping workbook {0}: not a valid .xlsx file ({1})".format(
                xlsx_path, e
            )
        )
        return False, stats
    except Exception as e:
        print(
            "WARNING: skipping workbook {0}: failed to load ({1})".format(
                xlsx_path, e
            )
        )
        return False, stats
    dirty = False

    for ws in wb.worksheets:
        stats.sheets += 1
        poss_col = _find_or_create_poss_cat_col(ws)

        for r in _iter_data_rows(ws):
            if limit is not None and stats.rows_updated >= limit:
                break

            stats.rows_seen += 1
            # Timed progress reporting
            now = time.time()
            if now - stats.last_progress_ts >= progress_interval:
                stats.last_progress_ts = now
                if recat_enabled:
                    print(
                        "  …progress {0}: rows_seen={1} updated={2} assigned={3} unassigned={4} rec_seen={5} rec_changed={6} rec_skipped={7}".format(
                            xlsx_path.name,
                            stats.rows_seen,
                            stats.rows_updated,
                            stats.rows_assigned,
                            stats.rows_unassigned,
                            stats.rows_recategorise_seen,
                            stats.rows_recategorise_changed,
                            stats.rows_recategorise_skipped,
                        )
                    )
                else:
                    print(
                        "  …progress {0}: rows_seen={1} updated={2} assigned={3} unassigned={4}".format(
                            xlsx_path.name,
                            stats.rows_seen,
                            stats.rows_updated,
                            stats.rows_assigned,
                            stats.rows_unassigned,
                        )
                    )

            # Step (5): optional recategorisation pass for existing poss_cat
            if recat_enabled:
                try:
                    existing_val_obj = ws.cell(row=r, column=poss_col).value
                except Exception:
                    existing_val_obj = None

                existing_list = _parse_poss_cat_cell(existing_val_obj)
                existing_list_l = [x.lower() for x in existing_list]

                consider = False
                if recat_scope == "all":
                    consider = bool(existing_list)
                else:
                    # unassigned scope: only rows whose cell is exactly 'unassigned'
                    consider = bool(existing_list_l) and ("unassigned" in existing_list_l) and (
                                len(existing_list_l) == 1)

                if consider:
                    stats.rows_recategorise_seen += 1
                    # Read row payload now (needed for LLM assignment)
                    row = _read_row_payload(ws, r)

                    # If the row is empty, treat as recategorise-skipped
                    if not (
                            _norm_cell_text(row.jyutping)
                            or _norm_cell_text(row.meaning)
                            or _norm_cell_text(row.hanzi)
                    ):
                        stats.rows_recategorise_skipped += 1
                        stats.rows_skipped_empty += 1
                        continue

                    # Ask assigner for categories (same rules as initial assignment)
                    proposed = _assign_for_row(
                        row,
                        allowed=allowed,
                        hanzi_to_cats=hanzi_to_cats,
                        assigner=assigner,
                        cache=cache,
                    )

                    merged = _merge_categories(existing_list, proposed, policy=recat_policy)
                    new_cell = _format_poss_cat_cell(merged)
                    old_cell = _format_poss_cat_cell(existing_list)

                    # --- Track recategorisation outcomes for reporting ---
                    new_list = _parse_poss_cat_cell(new_cell)
                    new_list_l = [x.lower() for x in new_list]
                    if new_list_l == ["unassigned"]:
                        stats.rows_recategorise_still_unassigned += 1
                        # Also collect for end-of-run unassigned analysis (so the report covers recategorisation results)
                        try:
                            stats.rows_unassigned += 1
                            wbname = workbook_label if workbook_label else xlsx_path.name
                            stats.unassigned_rows.append(
                                UnassignedRow(
                                    workbook=wbname,
                                    sheet=ws.title,
                                    row_index=r,
                                    jyutping=row.jyutping,
                                    meaning=row.meaning,
                                    hanzi=row.hanzi,
                                )
                            )
                        except Exception:
                            pass
                    else:
                        stats.rows_recategorise_became_assigned += 1
                        # Keep overall assigned count consistent with the normal path (used in per-file summary output)
                        try:
                            stats.rows_assigned += 1
                        except Exception:
                            pass

                    if new_cell and new_cell != old_cell:
                        stats.rows_recategorise_changed += 1
                        if not dry_run:
                            ws.cell(row=r, column=poss_col).value = new_cell
                            stats.rows_updated += 1
                            dirty = True
                    else:
                        stats.rows_recategorise_skipped += 1

                    # After recategorise path, skip the normal assignment flow to avoid double work.
                    continue

            # Normal assignment path reads row payload below
            row = _read_row_payload(ws, r)

            if not (_norm_cell_text(row.jyutping) or _norm_cell_text(row.meaning) or _norm_cell_text(row.hanzi)):
                stats.rows_skipped_empty += 1
                continue

            existing = _norm_cell_text(ws.cell(row=r, column=poss_col).value)
            if existing and not force:
                stats.rows_skipped_existing += 1

                # If we are proposing new categories, still collect rows that are explicitly
                # unassigned so we can analyse them without overwriting the sheet.
                if collect_unassigned_from_existing:
                    try:
                        parts = [p.strip() for p in str(existing).split(",") if p.strip()]
                    except Exception:
                        parts = [str(existing).strip()] if str(existing).strip() else []

                    parts_l = [p.lower() for p in parts]
                    # Treat as unassigned only if the cell is exactly "unassigned" (no other cats).
                    if parts_l and ("unassigned" in parts_l) and len(parts_l) == 1:
                        stats.rows_unassigned += 1
                        try:
                            wbname = workbook_label if workbook_label else xlsx_path.name
                            stats.unassigned_rows.append(
                                UnassignedRow(
                                    workbook=wbname,
                                    sheet=ws.title,
                                    row_index=r,
                                    jyutping=row.jyutping,
                                    meaning=row.meaning,
                                    hanzi=row.hanzi,
                                )
                            )
                        except Exception:
                            pass

                continue

            cats = _assign_for_row(
                row,
                allowed=allowed,
                hanzi_to_cats=hanzi_to_cats,
                assigner=assigner,
                cache=cache,
            )

            if not cats:
                cats = ["unassigned"]

            if "unassigned" in cats and len(cats) == 1:
                stats.rows_unassigned += 1
                # Record unassigned row for later reporting (defensively).
                try:
                    wbname = workbook_label if workbook_label else xlsx_path.name
                    stats.unassigned_rows.append(
                        UnassignedRow(
                            workbook=wbname,
                            sheet=ws.title,
                            row_index=r,
                            jyutping=row.jyutping,
                            meaning=row.meaning,
                            hanzi=row.hanzi,
                        )
                    )
                except Exception:
                    pass
            else:
                stats.rows_assigned += 1

            new_val = ", ".join(cats)
            if dry_run:
                continue

            ws.cell(row=r, column=poss_col).value = new_val
            stats.rows_updated += 1
            dirty = True

    if dry_run:
        return False, stats

    if not dirty:
        return False, stats

    if overwrite:
        out_path = xlsx_path
    else:
        out_path = xlsx_path.with_name("{0}.categorised{1}".format(xlsx_path.stem, xlsx_path.suffix))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return True, stats


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = _project_root()

    p = argparse.ArgumentParser(prog="add_category_data")
    p.add_argument(
        "--root",
        default=str(root),
        help="Project root (default: inferred from utils/add_category_data.py)",
    )
    p.add_argument(
        "--andy-dir",
        default="data/andy_ss",
        help="Directory (relative to root) containing .xlsx files",
    )
    p.add_argument(
        "--categories",
        default="data/categories.yaml",
        help="Categories YAML (relative to root)",
    )
    p.add_argument(
        "--cache",
        default="data/category_cache.yaml",
        help="Cache YAML (relative to root)",
    )
    p.add_argument("--dry-run", action="store_true", help="Do not write any files")
    p.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite input .xlsx files (default: write *.categorised.xlsx)",
    )
    p.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing poss_cat values",
    )
    p.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Stop after updating N rows (for testing)",
    )
    p.add_argument(
        "--provider",
        default="none",
        choices=["none", "openai"],
        help="Category assignment provider",
    )
    p.add_argument(
        "--model",
        default="gpt-4.1-mini",
        help="Model name for provider=openai",
    )
    p.add_argument(
        "--timeout-s",
        type=float,
        default=30.0,
        help="Per-request timeout (seconds) when provider=openai",
    )
    p.add_argument(
        "--max-retries",
        type=int,
        default=3,
        help="Max retries per row when provider=openai (not counting the first attempt)",
    )
    p.add_argument(
        "--backoff-s",
        type=float,
        default=1.0,
        help="Backoff base (seconds) between retries when provider=openai",
    )
    p.add_argument(
        "--throttle-s",
        type=float,
        default=0.0,
        help="Optional sleep (seconds) before each OpenAI request",
    )
    p.add_argument(
        "--preflight",
        action="store_true",
        help="When provider=openai, run a small preflight request before processing workbooks",
    )
    p.add_argument(
        "--propose-new-categories",
        action="store_true",
        help="After processing, propose candidate new categories based on unassigned rows (heuristic only)",
    )
    p.add_argument(
        "--propose-min-rows",
        type=int,
        default=3,
        help="Minimum unassigned rows required to propose a new category",
    )
    p.add_argument(
        "--propose-max",
        type=int,
        default=10,
        help="Maximum number of proposed categories to report",
    )

    p.add_argument(
        "--propose-provider",
        default="none",
        choices=["none", "openai"],
        help="Provider for proposing new categories (report only)",
    )
    p.add_argument(
        "--propose-model",
        default="gpt-4.1-mini",
        help="Model name for propose-provider=openai",
    )
    p.add_argument(
        "--propose-batch-size",
        type=int,
        default=25,
        help="Rows per proposal request when propose-provider=openai",
    )
    p.add_argument(
        "--export-proposed-categories",
        default="",
        help=(
            "When proposing new categories, write them to this YAML path (relative to root if not absolute). "
            "The file is categories.yaml-compatible so you can merge it using utils/expand_categories.py."
        ),
    )
    p.add_argument(
        "--export-include-example-hanzi",
        action="store_true",
        help="When exporting proposed categories, seed each category with example Hanzi from the report.",
    )

    # --- Option to apply proposed categories directly to categories.yaml ---
    p.add_argument(
        "--apply-proposed-categories",
        action="store_true",
        help="When proposing new categories, directly merge them into the categories YAML (preserving comments/order).",
    )
    p.add_argument(
        "--apply-proposed-seed-hanzi",
        action="store_true",
        help="When merging proposed categories, seed new categories with example Hanzi from the proposal report.",
    )

    p.add_argument(
        "--recategorise",
        action="store_true",
        help="Reconsider existing poss_cat values (uses provider if enabled).",
    )
    p.add_argument(
        "--recategorise-scope",
        default="unassigned",
        choices=["unassigned", "all"],
        help="When recategorising, which rows to consider: only 'unassigned' or 'all'.",
    )
    p.add_argument(
        "--recategorise-policy",
        default="append",
        choices=["append", "replace"],
        help="When recategorising, append to existing categories or replace them.",
    )

    args = p.parse_args(list(argv) if argv is not None else None)

    root_path = Path(str(args.root)).resolve()
    andy_dir = (root_path / str(args.andy_dir)).resolve()
    categories_path = (root_path / str(args.categories)).resolve()
    cache_path = (root_path / str(args.cache)).resolve()

    categories_map = _load_categories_map(categories_path)
    allowed_categories = _allowed_categories(categories_map)

    assigner = _choose_assigner(
        str(args.provider),
        str(args.model),
        timeout_s=float(args.timeout_s),
        max_retries=int(args.max_retries),
        backoff_s=float(args.backoff_s),
        throttle_s=float(args.throttle_s),
    )

    # Optional preflight for OpenAI: fail fast if connectivity/auth/model is broken.
    if _norm_cell_text(str(args.provider)).lower() == "openai" and bool(args.preflight):
        try:
            if hasattr(assigner, "preflight") and callable(getattr(assigner, "preflight")):
                ok, msg = assigner.preflight()  # type: ignore
                print(msg)
                if not ok:
                    return 2
        except Exception as e:
            print("OpenAI preflight failed: {0}".format(e))
            return 2
        print(
            "OpenAI assignment starting (model={0}, timeout={1}s, retries={2})".format(
                args.model, args.timeout_s, args.max_retries
            )
        )

    cache_raw = _load_yaml(cache_path)
    cache: Dict[str, object] = cache_raw if isinstance(cache_raw, dict) else {}

    if not andy_dir.exists():
        raise SystemExit("Andy directory not found: {0}".format(andy_dir))

    xlsx_files = sorted(list(andy_dir.rglob("*.xlsx")))
    if not xlsx_files:
        print("No .xlsx files found under {0}".format(andy_dir))
        return 0

    total = RunStats(files=0)
    all_unassigned: List[UnassignedRow] = []

    print("Loaded {0} categories (including unassigned).".format(len(allowed_categories)))
    print("Scanning {0} workbooks under {1}".format(len(xlsx_files), andy_dir))

    for fp in xlsx_files:
        changed, stats = process_workbook(
            fp,
            categories_map=categories_map,
            assigner=assigner,
            cache=cache,
            dry_run=bool(args.dry_run),
            overwrite=bool(args.overwrite),
            force=bool(args.force),
            limit=args.limit,
            workbook_label=str(fp.relative_to(root_path)),
            collect_unassigned_from_existing=bool(getattr(args, "propose_new_categories", False)),
            recategorise=bool(getattr(args, "recategorise", False)),
            recategorise_scope=str(getattr(args, "recategorise_scope", "unassigned")),
            recategorise_policy=str(getattr(args, "recategorise_policy", "append")),
        )

        total.files += 1
        total.sheets += stats.sheets
        total.rows_seen += stats.rows_seen
        total.rows_updated += stats.rows_updated
        total.rows_skipped_existing += stats.rows_skipped_existing
        total.rows_skipped_empty += stats.rows_skipped_empty
        total.rows_unassigned += stats.rows_unassigned
        total.rows_assigned += stats.rows_assigned
        # Accumulate recategorisation stats
        total.rows_recategorise_seen += stats.rows_recategorise_seen
        total.rows_recategorise_changed += stats.rows_recategorise_changed
        total.rows_recategorise_skipped += stats.rows_recategorise_skipped
        total.rows_recategorise_still_unassigned += stats.rows_recategorise_still_unassigned
        total.rows_recategorise_became_assigned += stats.rows_recategorise_became_assigned

        status = "CHANGED" if changed else "OK"
        print(
            "[{0}] {1} | sheets={2} rows_seen={3} updated={4} skipped_existing={5} empty={6} unassigned={7} assigned={8}".format(
                status,
                fp.relative_to(root_path),
                stats.sheets,
                stats.rows_seen,
                stats.rows_updated,
                stats.rows_skipped_existing,
                stats.rows_skipped_empty,
                stats.rows_unassigned,
                stats.rows_assigned,
            )
        )
        # Per-workbook PDF report (written next to the workbook)
        try:
            per_file_unassigned = list(getattr(stats, "unassigned_rows", []) or [])
        except Exception:
            per_file_unassigned = []
        try:
            _write_pdf_report_for_workbook(
                fp,
                root_path=root_path,
                stats=stats,
                unassigned_rows=per_file_unassigned,
                categories=allowed_categories,
                dry_run=bool(args.dry_run),
            )
        except Exception as e:
            print("WARNING: failed to write PDF report for {0}: {1}".format(fp, e))

        # Collect unassigned rows for later category proposal
        try:
            if getattr(stats, "unassigned_rows", None):
                all_unassigned.extend(stats.unassigned_rows)
        except Exception:
            pass

    # Persist cache
    if not bool(args.dry_run):
        try:
            _save_yaml(cache_path, cache)
        except Exception as e:
            print("WARNING: failed to write cache: {0}".format(e))

    print(
        "Done. files={0} sheets={1} rows_seen={2} updated={3} skipped_existing={4} empty={5} unassigned={6} assigned={7}".format(
            total.files,
            total.sheets,
            total.rows_seen,
            total.rows_updated,
            total.rows_skipped_existing,
            total.rows_skipped_empty,
            total.rows_unassigned,
            total.rows_assigned,
        )
    )

    # --- Recategorisation summary ---
    print(
        "Recategorisation summary: rows_recategorise_seen={0} rows_recategorise_changed={1} rows_recategorise_skipped={2} still_unassigned={3} became_assigned={4}".format(
            total.rows_recategorise_seen,
            total.rows_recategorise_changed,
            total.rows_recategorise_skipped,
            total.rows_recategorise_still_unassigned,
            total.rows_recategorise_became_assigned,
        )
    )

    # --- End-of-run unassigned analysis report ---
    print()
    if all_unassigned:
        _print_unassigned_analysis(all_unassigned, max_rows=None)
    else:
        # If recategorisation ran and still left rows unassigned, make that explicit.
        if total.rows_recategorise_still_unassigned:
            print(
                "Unassigned analysis: no unassigned rows were collected, but recategorisation still_unassigned={0}.".format(
                    total.rows_recategorise_still_unassigned
                )
            )
            print(
                "Tip: re-run with --force (or run without --recategorise) if you want the normal assignment path to collect these rows."
            )
        else:
            print("Unassigned analysis: no unassigned rows remain.")

    # --- Propose new categories (heuristic report) ---
    if getattr(args, "propose_new_categories", False):
        print()
        print("New category proposal (unassigned rows):")
        print("  Found {0} unassigned rows across all workbooks.".format(len(all_unassigned)))

        proposer = _choose_category_proposer(
            str(getattr(args, "propose_provider", "none")),
            str(getattr(args, "propose_model", "gpt-4.1-mini")),
            timeout_s=float(args.timeout_s),
            max_retries=max(0, int(getattr(args, "max_retries", 0))),
            backoff_s=float(getattr(args, "backoff_s", 1.0)),
            throttle_s=float(getattr(args, "throttle_s", 0.0)),
        )

        proposals: List[Dict[str, object]] = []
        if proposer is None:
            print("  propose-provider=none: no LLM proposal step (report will be empty).")
        else:
            if bool(getattr(args, "preflight", False)):
                try:
                    if hasattr(proposer, "preflight") and callable(getattr(proposer, "preflight")):
                        ok, msg = proposer.preflight()  # type: ignore
                        print("  " + str(msg))
                        if not ok:
                            print("  Proposer preflight failed; skipping proposals.")
                            proposer = None
                except Exception as e:
                    print("  Proposer preflight failed: {0}".format(e))
                    proposer = None

            if proposer is not None:
                proposals = proposer.propose(
                    all_unassigned,
                    existing_categories=allowed_categories,
                    min_rows=int(getattr(args, "propose_min_rows", 3)),
                    max_categories=int(getattr(args, "propose_max", 10)),
                    batch_size=int(getattr(args, "propose_batch_size", 25)),
                )

        _print_proposed_categories_report(proposals, prefix="  ")

        # --- Category proposal summary ---
        print("Category proposal summary: proposals_generated={0}".format(len(proposals)))

        export_path_raw = _norm_cell_text(getattr(args, "export_proposed_categories", ""))
        if export_path_raw:
            try:
                export_path = Path(export_path_raw)
                if not export_path.is_absolute():
                    export_path = (root_path / export_path_raw).resolve()
                _export_proposed_categories_yaml(
                    proposals,
                    export_path,
                    include_example_hanzi=bool(getattr(args, "export_include_example_hanzi", False)),
                )
                print("  Exported proposed categories to: {0}".format(export_path))
                print("  Tip: merge this into data/categories.yaml using utils/expand_categories.py")
            except Exception as e:
                print("  WARNING: failed to export proposed categories: {0}".format(e))

        # Optionally apply proposals directly to categories.yaml
        apply_flag = bool(getattr(args, "apply_proposed_categories", False))
        if apply_flag:
            if proposals:
                try:
                    added, skipped = _merge_proposed_categories_into_categories_yaml(
                        proposals,
                        categories_path,
                        seed_example_hanzi=bool(getattr(args, "apply_proposed_seed_hanzi", False)),
                    )
                    print("  Merged proposed categories into: {0}".format(categories_path))
                    print("  Added {0} new categories (skipped {1} existing).".format(added, skipped))
                    print(
                        "Category proposal summary: proposals_generated={0} proposals_applied={1} skipped_existing={2}".format(
                            len(proposals), added, skipped
                        ))
                    print()
                    _print_unassigned_pattern_summary_from_proposals(proposals, all_unassigned)
                except Exception as e:
                    print("  WARNING: failed to merge proposed categories: {0}".format(e))
            else:
                print("  No proposals to apply.")


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
